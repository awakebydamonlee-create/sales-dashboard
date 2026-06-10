#!/usr/bin/env python3
"""일일 출고 모니터링 보고서 — 배드블루 출고표 일일 분석.

사용법:
    python3 daily_inventory_report.py <xlsx_path>

출력:
    - inventory_daily_report.html
    - daily_report_summary.json (채팅 알림용)
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

# 컬럼 인덱스 (0-based)
COL_CODE = 0
COL_BIGO = 1
COL_PRODUCT = 2
COL_SIZE = 3
COL_STOPPED = 9     # 판매중단
COL_STOCK = 10      # 물류재고
COL_INCOMING = 11   # 입고예정
COL_REAL_STOCK = 12 # 예약배송 출고 후 실재고
COL_OUTBOUND = 13   # 출고합계+

# 분석 파라미터
AVG_DAYS = 14       # 일평균 산출 기간 (영업일 기준)
URGENT_DAYS = 7     # 발주 우선 기준 (잔여일)
WARN_DAYS = 14      # 모니터 기준 (잔여일)
TARGET_DAYS = 30    # 발주 보충 기준 (30일치 확보)
SPIKE_RATIO = 1.5   # 갑자기 잘나감 기준


def to_float(v):
    if v is None or v == '' or v == '#REF!':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def is_stopped(v):
    return v is not None and str(v).strip().upper() in ('O', 'X', 'TRUE', '1', '판매중단', '중단')


def parse_size(v):
    if v is None or v == '':
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def product_label(name, size):
    name = (name or '').strip()
    size = parse_size(size)
    if not size or size.lower() == 'free':
        return name
    return f"{name} <small style='color:#888;'>SIZE {size}</small>"


def parse_sheet_to_dict(ws):
    """시트를 dict[code] -> row dict 로 변환."""
    result = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[COL_CODE]:
            continue
        code = str(row[COL_CODE]).strip()
        if not code or code == 'CODE':
            continue
        # 같은 CODE는 시트당 한 행만 존재 (size 별로 다른 CODE 사용)
        result[code] = {
            'code': code,
            'bigo': row[COL_BIGO] if len(row) > COL_BIGO else None,
            'product': row[COL_PRODUCT] if len(row) > COL_PRODUCT else None,
            'size': row[COL_SIZE] if len(row) > COL_SIZE else None,
            'stopped': is_stopped(row[COL_STOPPED] if len(row) > COL_STOPPED else None),
            'stock': to_float(row[COL_STOCK] if len(row) > COL_STOCK else None),
            'incoming': to_float(row[COL_INCOMING] if len(row) > COL_INCOMING else None),
            'real_stock': to_float(row[COL_REAL_STOCK] if len(row) > COL_REAL_STOCK else None),
            'outbound': to_float(row[COL_OUTBOUND] if len(row) > COL_OUTBOUND else None),
        }
    return result


def parse_incoming_dates(bigo):
    """비고에서 입고 예정 날짜를 파싱."""
    if not bigo:
        return []
    text = str(bigo).strip()
    # 패턴: "예약배송 5/30", "예약배송 5/22 2차 6/12", "추후 설정 (입고 예정 5/28)"
    dates = re.findall(r'(\d{1,2})/(\d{1,2})', text)
    return [f"{m}/{d}" for m, d in dates]


def has_incoming_keyword(bigo):
    if not bigo:
        return False
    s = str(bigo)
    return ('예약배송' in s) or ('입고 예정' in s) or ('입고예정' in s)


def main(xlsx_path):
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    # 날짜 형식 시트만 필터 (예: 5.14)
    date_sheets = [s for s in sheet_names if re.fullmatch(r'\d{1,2}\.\d{1,2}', s)]

    # 가장 최근 시트 = 오늘 (5.15) → 베이스라인으로 안 씀 (오늘은 진행중)
    # 두 번째 = 어제 close (5.14)
    if len(date_sheets) < 2:
        raise RuntimeError(f"날짜 시트가 부족합니다: {date_sheets}")

    today_sheet = date_sheets[0]   # 5.15 (진행중)
    yday_sheet = date_sheets[1]    # 5.14 close
    prev_sheet = date_sheets[2] if len(date_sheets) > 2 else None  # 5.13

    print(f"오늘 시트: {today_sheet} / 어제 close: {yday_sheet} / 그저께: {prev_sheet}", file=sys.stderr)

    # 어제 close 데이터 로드
    yday_data = parse_sheet_to_dict(wb[yday_sheet])
    prev_data = parse_sheet_to_dict(wb[prev_sheet]) if prev_sheet else {}

    # 일평균 산출 — 최근 AVG_DAYS 영업일 (오늘 제외, 어제 포함)
    recent_sheets = date_sheets[1:1 + AVG_DAYS]
    print(f"일평균 산출 대상 시트({len(recent_sheets)}개): {recent_sheets}", file=sys.stderr)

    outbound_history = defaultdict(list)  # code -> [outbound_qty, ...]
    for sn in recent_sheets:
        try:
            ws = wb[sn]
            sheet_data = parse_sheet_to_dict(ws)
            for code, row in sheet_data.items():
                outbound_history[code].append(row['outbound'])
        except Exception as e:
            print(f"  Skip {sn}: {e}", file=sys.stderr)

    avg_outbound = {}  # code -> 일평균
    for code, vals in outbound_history.items():
        if not vals:
            avg_outbound[code] = 0.0
        else:
            avg_outbound[code] = sum(vals) / len(vals)

    # 일평균 (어제 제외) — 갑자기 잘나감 비교용
    avg_outbound_exyd = {}
    for code, vals in outbound_history.items():
        prior = vals[1:] if len(vals) > 1 else vals
        avg_outbound_exyd[code] = sum(prior) / len(prior) if prior else 0.0

    # ---- 섹션 분석 ----
    # 핵심 룰: 입고예정 있으면 어차피 채워지니까 발주 카테고리에서 제외
    urgent = []     # 발주 우선 = 입고예정 0 + 잔여 ≤ 7일
    refill = []     # 발주 보충 = 입고예정 0 + 잔여 7~30일
    yd_top = []     # 어제 출고 TOP (모든 출고 양수, 정렬해서 잘라냄)
    spikes = []     # 갑자기 잘나감
    yd_in = []      # 어제 입고
    incoming_planned = []  # 입고 예정
    zero_stock = [] # 재고 0 SKU (판매중단 제외)

    for code, row in yday_data.items():
        if row['stopped']:
            continue
        avg = avg_outbound.get(code, 0.0)
        real_stock = row['real_stock']
        stock = row['stock']
        incoming = row['incoming']
        product = row['product'] or ''
        size = row['size']
        bigo = row['bigo']

        days_left = (real_stock / avg) if avg > 0 else float('inf')
        no_incoming = (incoming is None or incoming == 0)

        # 발주 우선: 입고예정 0 AND 잔여 ≤ 7일
        if avg > 0 and no_incoming and days_left <= URGENT_DAYS:
            urgent.append({
                'code': code, 'product': product, 'size': size,
                'real_stock': int(real_stock), 'avg': avg, 'days_left': days_left,
            })

        # 발주 보충: 입고예정 0 AND 잔여 7~30일 (일평균 0.5 이상 노이즈 제거)
        elif avg >= 0.5 and no_incoming and URGENT_DAYS < days_left <= TARGET_DAYS:
            refill.append({
                'code': code, 'product': product, 'size': size,
                'real_stock': int(real_stock), 'avg': avg, 'days_left': days_left,
            })

        # 어제 출고
        if row['outbound'] > 0:
            yd_top.append({
                'code': code, 'product': product, 'size': size,
                'outbound': int(row['outbound']), 'real_stock': int(real_stock),
            })
            # 갑자기 잘나감 — 평소 = 14일 일평균 (어제 포함)
            if avg > 0 and row['outbound'] >= avg * SPIKE_RATIO and row['outbound'] >= 2:
                spikes.append({
                    'code': code, 'product': product, 'size': size,
                    'outbound': int(row['outbound']), 'avg': avg,
                    'ratio': row['outbound'] / avg,
                })

        # 어제 입고: 그저께 vs 어제 stock 비교, 입고된 수량 = (어제 stock - 그저께 stock) + 어제 출고
        if prev_data and code in prev_data:
            prev_stock = prev_data[code]['stock']
            inbound = (stock - prev_stock) + row['outbound']
            if inbound > 0.5:
                yd_in.append({
                    'code': code, 'product': product, 'size': size,
                    'inbound': int(round(inbound)), 'stock': int(stock),
                })

        # 입고 예정 (비고에 날짜)
        if has_incoming_keyword(bigo) and incoming and incoming > 0:
            dates = parse_incoming_dates(bigo)
            incoming_planned.append({
                'code': code, 'product': product, 'size': size,
                'qty': int(incoming),
                'schedule': ' / '.join(dates) if dates else '-',
                'bigo': str(bigo).strip(),
            })

        # 재고 0
        if real_stock == 0 and stock == 0:
            zero_stock.append({
                'code': code, 'product': product, 'size': size,
                'avg': avg,
            })

    # 정렬
    urgent.sort(key=lambda x: x['days_left'])
    refill.sort(key=lambda x: x['days_left'])
    yd_top.sort(key=lambda x: -x['outbound'])
    spikes.sort(key=lambda x: -x['ratio'])
    yd_in.sort(key=lambda x: -x['inbound'])
    incoming_planned.sort(key=lambda x: x['schedule'])
    zero_stock.sort(key=lambda x: -x['avg'])

    # ---- HTML 생성 ----
    today_str = datetime.now().strftime('%Y-%m-%d')
    gen_time = datetime.now().strftime('%Y-%m-%d %H:%M')
    yday_label = yday_sheet  # e.g., 5.14

    def html_table(rows, headers, fmt_row):
        if not rows:
            return '<div class="empty">해당 항목 없음</div>'
        head_html = ''.join(f'<th>{h}</th>' for h in headers)
        body_html = ''.join(fmt_row(r) for r in rows)
        return f'<table><thead><tr>{head_html}</tr></thead><tbody>{body_html}</tbody></table>'

    # 섹션별 행 포맷
    def fmt_urgent(r):
        return f'<tr><td>{product_label(r["product"], r["size"])}</td><td class="">{parse_size(r["size"]) or "Free"}</td><td class="num">{r["real_stock"]}</td><td class="num">{r["avg"]:.1f}</td><td class="num urgent">{r["days_left"]:.1f}</td><td class="">{r["code"]}</td></tr>'

    def fmt_refill(r):
        return f'<tr><td>{product_label(r["product"], r["size"])}</td><td class="num">{r["real_stock"]}</td><td class="num">{r["avg"]:.1f}</td><td class="num warn">{r["days_left"]:.1f}</td><td class="">{r["code"]}</td></tr>'

    def fmt_yd_top(r):
        return f'<tr><td>{product_label(r["product"], r["size"])}</td><td class="num">{r["outbound"]}</td><td class="num">{r["real_stock"]}</td></tr>'

    def fmt_spike(r):
        return f'<tr><td>{product_label(r["product"], r["size"])}</td><td class="num">{r["outbound"]}</td><td class="num">{r["avg"]:.1f}</td><td class="num">{r["ratio"]:.1f}배</td></tr>'

    def fmt_yd_in(r):
        return f'<tr><td>{product_label(r["product"], r["size"])}</td><td class="num">{r["inbound"]}</td><td class="num">{r["stock"]}</td><td class="">{r["code"]}</td></tr>'

    def fmt_incoming(r):
        return f'<tr><td>{product_label(r["product"], r["size"])}</td><td class="num">{r["qty"]}</td><td>{r["schedule"]}</td><td class="">{r["bigo"]}</td></tr>'

    def fmt_zero(r):
        return f'<tr><td>{product_label(r["product"], r["size"])}</td><td class="num">{r["avg"]:.1f}</td><td class="">{r["code"]}</td></tr>'

    yd_top_show = yd_top[:15]

    sections_html = ''
    sections_html += f'<div class="section red" id="sec-red"><h2>🔴 발주 우선 <span style="color:#999;font-weight:500;font-size:14px;">({len(urgent)}개)</span></h2><p class="desc">잔여 {URGENT_DAYS}일 이내, 입고예정 0</p>'
    sections_html += html_table(urgent, ['상품명', 'SIZE', '실재고', '일평균', '잔여일', 'CODE'], fmt_urgent) + '</div>'

    sections_html += f'<div class="section orange" id="sec-orange"><h2>🟠 발주 보충 <span style="color:#999;font-weight:500;font-size:14px;">({len(refill)}개)</span></h2><p class="desc">입고예정 0 + 잔여 {URGENT_DAYS}~{TARGET_DAYS}일</p>'
    sections_html += html_table(refill, ['상품명', '실재고', '일평균', '잔여일', 'CODE'], fmt_refill) + '</div>'

    sections_html += f'<div class="section blue" id="sec-blue"><h2>📊 어제 출고 TOP10 <span style="color:#999;font-weight:500;font-size:14px;">({len(yd_top_show)}개)</span></h2><p class="desc">하루 출고량 상위</p>'
    sections_html += html_table(yd_top_show, ['상품명', '출고', '실재고_후'], fmt_yd_top) + '</div>'

    sections_html += f'<div class="section purple" id="sec-purple"><h2>🚀 갑자기 잘나감 <span style="color:#999;font-weight:500;font-size:14px;">({len(spikes)}개)</span></h2><p class="desc">평소 대비 {SPIKE_RATIO}배 이상</p>'
    sections_html += html_table(spikes, ['상품명', '어제 출고', '평소', '비율'], fmt_spike) + '</div>'

    sections_html += f'<div class="section green" id="sec-green"><h2>📦 어제 입고 <span style="color:#999;font-weight:500;font-size:14px;">({len(yd_in)}개)</span></h2><p class="desc">입고 발생</p>'
    sections_html += html_table(yd_in, ['상품명', '입고수량', '재고_후', 'CODE'], fmt_yd_in) + '</div>'

    sections_html += f'<div class="section cyan" id="sec-cyan"><h2>📅 입고 예정 <span style="color:#999;font-weight:500;font-size:14px;">({len(incoming_planned)}개)</span></h2><p class="desc">비고에 등록된 예약배송</p>'
    sections_html += html_table(incoming_planned, ['상품명', '수량', '일정', '비고'], fmt_incoming) + '</div>'

    sections_html += f'<div class="section gray" id="sec-gray"><h2>⚫ 재고 0 SKU <span style="color:#999;font-weight:500;font-size:14px;">({len(zero_stock)}개)</span></h2><p class="desc">판매중단 제외</p>'
    sections_html += html_table(zero_stock, ['상품명', '일평균', 'CODE'], fmt_zero) + '</div>'

    # Summary cards
    cards = [
        ('🔴', len(urgent), '발주 우선', 'red'),
        ('🟠', len(refill), '발주 보충', 'orange'),
        ('📊', len(yd_top_show), '어제 출고 TOP10', 'blue'),
        ('🚀', len(spikes), '갑자기 잘나감', 'purple'),
        ('📦', len(yd_in), '어제 입고', 'green'),
        ('📅', len(incoming_planned), '입고 예정', 'cyan'),
        ('⚫', len(zero_stock), '재고 0 SKU', 'gray'),
    ]
    cards_html = ''
    for icon, count, label, color in cards:
        cards_html += f'<div class="sum-card" onclick="document.getElementById(\'sec-{color}\').scrollIntoView({{behavior:\'smooth\'}})"><div class="sum-icon">{icon}</div><div class="sum-count">{count}</div><div class="sum-label">{label}</div></div>'

    html = f'''<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<title>일일 출고 모니터링 보고서 — {today_str}</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Pretendard', sans-serif; background: #fafafa; padding: 20px; color: #1a1a1a; }}
.container {{ max-width: 1300px; margin: 0 auto; }}
h1 {{ font-size: 24px; margin-bottom: 4px; }}
.subtitle {{ color: #666; font-size: 13px; margin-bottom: 20px; }}
.summary {{ display: grid; grid-template-columns: repeat(7, 1fr); gap: 10px; margin-bottom: 24px; }}
.sum-card {{ background: white; border-radius: 10px; padding: 14px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); cursor: pointer; transition: transform 0.15s; }}
.sum-card:hover {{ transform: translateY(-2px); }}
.sum-icon {{ font-size: 18px; }}
.sum-count {{ font-size: 22px; font-weight: 700; margin-top: 4px; }}
.sum-label {{ font-size: 11px; color: #666; margin-top: 2px; }}
.section {{ background: white; border-radius: 12px; padding: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); margin-bottom: 16px; }}
.section h2 {{ font-size: 16px; margin-bottom: 4px; }}
.section .desc {{ color: #888; font-size: 12px; margin-bottom: 12px; }}
.section.red h2 {{ color: #e74c3c; }} .section.orange h2 {{ color: #ff8c42; }}
.section.yellow h2 {{ color: #d4a017; }} .section.blue h2 {{ color: #4a90e2; }}
.section.purple h2 {{ color: #8e44ad; }} .section.green h2 {{ color: #27ae60; }}
.section.cyan h2 {{ color: #16a085; }} .section.gray h2 {{ color: #888; }}
table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
th {{ background: #f8f8f8; padding: 8px 10px; text-align: left; font-weight: 600; color: #555; font-size: 11px; text-transform: uppercase; letter-spacing: 0.3px; }}
td {{ padding: 8px 10px; border-top: 1px solid #f0f0f0; }}
td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
td.urgent {{ color: #e74c3c; font-weight: 700; }}
td.warn {{ color: #ff8c42; font-weight: 600; }}
.empty {{ color: #aaa; text-align: center; padding: 20px; font-size: 13px; }}
</style></head><body><div class="container">
<h1>📅 일일 출고 모니터링 보고서</h1>
<p class="subtitle">기준 출고표: {yday_label} close · 생성: {gen_time}</p>
<div class="summary">{cards_html}</div>{sections_html}</div></body></html>'''

    out_html = xlsx_path.parent / 'inventory_daily_report.html'
    out_html.write_text(html, encoding='utf-8')
    print(f"HTML 저장: {out_html}", file=sys.stderr)

    # ---- 요약 JSON (채팅 알림용) ----
    summary = {
        'date': today_str,
        'baseline': f'{yday_label} close',
        'urgent_count': len(urgent),
        'urgent_top': [r['product'] for r in urgent[:3]],
        'refill_count': len(refill),
        'refill_top': [{'product': r['product'], 'days_left': round(r['days_left'], 1)} for r in refill[:3]],
        'spike_count': len(spikes),
        'spike_top': [{'product': r['product'], 'ratio': round(r['ratio'], 1)} for r in spikes[:2]],
        'incoming_planned_count': len(incoming_planned),
        'yd_in_count': len(yd_in),
        'yd_top_count': len(yd_top_show),
        'zero_stock_count': len(zero_stock),
    }

    out_json = xlsx_path.parent / 'daily_report_summary.json'
    out_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"JSON 저장: {out_json}", file=sys.stderr)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: daily_inventory_report.py <xlsx_path>", file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])
